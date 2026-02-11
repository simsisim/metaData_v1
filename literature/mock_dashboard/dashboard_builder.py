#!/usr/bin/env python3
"""
Excel Dashboard Builder for Trading System
==========================================

Creates professional Excel dashboards with conditional formatting,
charts, and institutional-grade presentation.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import json
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

class TradingDashboardBuilder:
    """Professional Excel dashboard builder for trading system"""
    
    def __init__(self, data_dir="sample_data", output_dir="output"):
        self.data_dir = Path(data_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # Define color scheme (institutional trading colors)
        self.colors = {
            'bullish': 'FF90EE90',      # Light Green
            'bearish': 'FFFF6B6B',      # Light Red  
            'neutral': 'FFFFEB9C',      # Light Yellow
            'strong_bullish': 'FF32CD32', # Lime Green
            'strong_bearish': 'FFDC143C', # Crimson
            'header': 'FF2F4F4F',       # Dark Slate Gray
            'subheader': 'FF708090',    # Slate Gray
            'background': 'FFF8F8FF'    # Ghost White
        }
        
        # Define fonts
        self.fonts = {
            'header': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
            'subheader': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'data': Font(name='Calibri', size=10),
            'signal': Font(name='Calibri', size=11, bold=True),
            'title': Font(name='Calibri', size=16, bold=True)
        }
        
    def load_mock_data(self):
        """Load all mock data files"""
        try:
            # Load market pulse data
            with open(self.data_dir / 'market_pulse_data.json', 'r') as f:
                market_pulse = json.load(f)
            
            # Load CSV files
            screener_results = pd.read_csv(self.data_dir / 'screener_results.csv')
            sector_performance = pd.read_csv(self.data_dir / 'sector_performance.csv')
            index_technical = pd.read_csv(self.data_dir / 'index_technical.csv')
            alerts = pd.read_csv(self.data_dir / 'alerts_data.csv')
            opportunities = pd.read_csv(self.data_dir / 'top_opportunities.csv')
            
            return {
                'market_pulse': market_pulse,
                'screener_results': screener_results,
                'sector_performance': sector_performance,
                'index_technical': index_technical,
                'alerts': alerts,
                'opportunities': opportunities
            }
            
        except Exception as e:
            print(f"❌ Error loading mock data: {e}")
            return None
    
    def create_market_pulse_tab(self, ws, data):
        """Create the Market Pulse overview tab"""
        
        # Set tab properties
        ws.title = "Market Pulse"
        
        # Main title
        ws['A1'] = "🚀 MARKET PULSE DASHBOARD"
        ws['A1'].font = self.fonts['title']
        ws.merge_cells('A1:J1')
        
        # Date and timestamp
        ws['A2'] = f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = self.fonts['data']
        
        # Market Status Section (A4:D8)
        ws['A4'] = "MARKET STATUS"
        ws['A4'].font = self.fonts['header']
        ws['A4'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('A4:D4')
        
        market_signal = data['market_pulse']['overall_signal']
        signal_emoji = {'BULLISH': '🟢', 'BEARISH': '🔴', 'NEUTRAL': '🟡'}[market_signal]
        
        ws['A5'] = f"Overall Signal: {signal_emoji} {market_signal}"
        ws['A5'].font = self.fonts['signal']
        
        confidence = data['market_pulse']['confidence']
        ws['A6'] = f"Confidence Level: {confidence}"
        
        # GMI Analysis (F4:I8)
        ws['F4'] = "GMI ANALYSIS"
        ws['F4'].font = self.fonts['header']
        ws['F4'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('F4:I4')
        
        gmi_data = data['market_pulse']['gmi_analysis']
        row = 5
        for index, gmi_info in gmi_data.items():
            signal_color = {'GREEN': self.colors['bullish'], 'RED': self.colors['bearish'], 'NEUTRAL': self.colors['neutral']}
            ws[f'F{row}'] = f"{index}: {gmi_info['gmi_signal']} ({gmi_info['gmi_score']}/4)"
            ws[f'F{row}'].fill = PatternFill(start_color=signal_color[gmi_info['gmi_signal']], 
                                           end_color=signal_color[gmi_info['gmi_signal']], fill_type='solid')
            row += 1
        
        # Risk Monitor (A10:D15)
        ws['A10'] = "RISK MONITOR"
        ws['A10'].font = self.fonts['header']
        ws['A10'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('A10:D10')
        
        dd_data = data['market_pulse']['distribution_days']
        ws['A11'] = f"Distribution Days: {dd_data['count']}/25"
        ws['A12'] = f"Warning Level: {dd_data['warning_level']}"
        
        # Color code warning level
        warning_colors = {'NONE': self.colors['bullish'], 'CAUTION': self.colors['neutral'], 'SEVERE': self.colors['bearish']}
        ws['A12'].fill = PatternFill(start_color=warning_colors[dd_data['warning_level']], 
                                    end_color=warning_colors[dd_data['warning_level']], fill_type='solid')
        
        # Market Breadth (F10:I15)
        ws['F10'] = "MARKET BREADTH"
        ws['F10'].font = self.fonts['header']
        ws['F10'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('F10:I10')
        
        breadth = data['market_pulse']['market_breadth']
        ws['F11'] = f"New Highs: {breadth['new_highs']}"
        ws['F12'] = f"New Lows: {breadth['new_lows']}"
        ws['F13'] = f"Net H/L: {breadth['net_highs']}"
        ws['F14'] = f"Status: {breadth['breadth_signal']}"
        
        # Color code breadth status
        breadth_colors = {'HEALTHY': self.colors['bullish'], 'UNHEALTHY': self.colors['bearish'], 'NEUTRAL': self.colors['neutral']}
        ws['F14'].fill = PatternFill(start_color=breadth_colors[breadth['breadth_signal']], 
                                    end_color=breadth_colors[breadth['breadth_signal']], fill_type='solid')
        
        # Sector Performance Table (A17:E26)
        ws['A17'] = "SECTOR PERFORMANCE (TODAY)"
        ws['A17'].font = self.fonts['header']
        ws['A17'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('A17:E17')
        
        # Headers
        headers = ['ETF', 'Sector', 'Daily %', 'Volume', 'Trend']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=18, column=i, value=header)
            cell.font = self.fonts['subheader']
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        
        # Sector data
        sector_df = data['sector_performance']
        for i, (_, row) in enumerate(sector_df.iterrows(), 19):
            ws[f'A{i}'] = row['etf']
            ws[f'B{i}'] = row['sector_name']
            ws[f'C{i}'] = f"{row['daily_change_pct']:+.1f}%"
            ws[f'D{i}'] = f"{row['volume']:,.0f}"
            ws[f'E{i}'] = row['trend_status']
            
            # Color code performance
            perf_color = self.colors['bullish'] if row['daily_change_pct'] > 0 else self.colors['bearish']
            ws[f'C{i}'].fill = PatternFill(start_color=perf_color, end_color=perf_color, fill_type='solid')
        
        # Auto-adjust column widths
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    max_length = max(max_length, len(str(cell.value)))
            
            ws.column_dimensions[column_letter].width = min(max_length + 2, 20)
    
    def create_screener_heatmap_tab(self, ws, data):
        """Create the Screener Heatmap tab"""
        
        ws.title = "Screener Heatmap"
        
        # Title
        ws['A1'] = "🔥 SCREENER OPPORTUNITIES HEATMAP"
        ws['A1'].font = self.fonts['title']
        ws.merge_cells('A1:H1')
        
        # Top Opportunities Table
        ws['A3'] = "TOP OPPORTUNITIES (Signal Strength Ranked)"
        ws['A3'].font = self.fonts['header']
        ws['A3'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('A3:H3')
        
        # Headers
        headers = ['Ticker', 'Signal', 'Primary Screener', 'Score', 'Price', 'Volume', 'Setup', 'R:R']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i, value=header)
            cell.font = self.fonts['subheader']
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        
        # Opportunities data
        opportunities_df = data['opportunities'].head(15)  # Top 15
        for i, (_, row) in enumerate(opportunities_df.iterrows(), 5):
            ws[f'A{i}'] = row['ticker']
            ws[f'B{i}'] = row['signal_strength']
            ws[f'C{i}'] = row['primary_screener'].replace('_', ' ').title()
            ws[f'D{i}'] = row['score']
            ws[f'E{i}'] = f"${row['current_price']:.2f}"
            ws[f'F{i}'] = f"{row['volume']:,.0f}"
            ws[f'G{i}'] = row['setup_stage']
            ws[f'H{i}'] = f"{row['risk_reward']:.1f}:1"
            
            # Color code signal strength
            signal_colors = {'STRONG': self.colors['strong_bullish'], 'MODERATE': self.colors['neutral'], 'WEAK': self.colors['bearish']}
            ws[f'B{i}'].fill = PatternFill(start_color=signal_colors[row['signal_strength']], 
                                         end_color=signal_colors[row['signal_strength']], fill_type='solid')
        
        # Screener Summary (A25:D35)
        ws['A25'] = "SCREENER SUITE SUMMARY"
        ws['A25'].font = self.fonts['header']
        ws['A25'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws.merge_cells('A25:D25')
        
        # Mock screener hit counts
        screener_summary = [
            ('Stockbee Suite', 12, 'STRONG'),
            ('Qullamaggie Suite', 6, 'MODERATE'),
            ('Volume Suite', 8, 'STRONG'),
            ('Gold Launch Pad', 3, 'WEAK'),
            ('RTI Screener', 7, 'MODERATE'),
            ('ADL Screener', 4, 'WEAK'),
            ('Guppy GMMA', 5, 'MODERATE'),
            ('ATR1 Suite', 9, 'STRONG')
        ]
        
        # Headers
        ws['A26'] = "Screener"
        ws['B26'] = "Hits"
        ws['C26'] = "Status"
        ws['D26'] = "Signal"
        
        for col in ['A26', 'B26', 'C26', 'D26']:
            ws[col].font = self.fonts['subheader']
            ws[col].fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        
        # Data
        for i, (screener, hits, status) in enumerate(screener_summary, 27):
            ws[f'A{i}'] = screener
            ws[f'B{i}'] = hits
            ws[f'C{i}'] = status
            
            # Signal indicator
            if status == 'STRONG':
                ws[f'D{i}'] = '🟢'
                status_color = self.colors['bullish']
            elif status == 'MODERATE':
                ws[f'D{i}'] = '🟡'
                status_color = self.colors['neutral']
            else:
                ws[f'D{i}'] = '🔴'
                status_color = self.colors['bearish']
            
            ws[f'C{i}'].fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
    
    def create_alerts_tab(self, ws, data):
        """Create the Alerts & Actions tab"""
        
        ws.title = "Alerts & Actions"
        
        # Title
        ws['A1'] = "🚨 MARKET ALERTS & ACTION ITEMS"
        ws['A1'].font = self.fonts['title']
        ws.merge_cells('A1:F1')
        
        # High Priority Alerts
        ws['A3'] = "🚨 HIGH PRIORITY ALERTS"
        ws['A3'].font = self.fonts['header']
        ws['A3'].fill = PatternFill(start_color=self.colors['strong_bearish'], end_color=self.colors['strong_bearish'], fill_type='solid')
        ws.merge_cells('A3:F3')
        
        high_alerts = data['alerts'][data['alerts']['priority'] == 'HIGH']
        row = 4
        for _, alert in high_alerts.iterrows():
            ws[f'A{row}'] = f"• {alert['message']}"
            ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
            row += 1
        
        # Medium Priority Section
        row += 2
        ws[f'A{row}'] = "⚠️ MEDIUM PRIORITY ALERTS"
        ws[f'A{row}'].font = self.fonts['header'] 
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['neutral'], end_color=self.colors['neutral'], fill_type='solid')
        ws.merge_cells(f'A{row}:F{row}')
        
        medium_alerts = data['alerts'][data['alerts']['priority'] == 'MEDIUM']
        row += 1
        for _, alert in medium_alerts.iterrows():
            ws[f'A{row}'] = f"• {alert['message']}"
            row += 1
        
        # Action Items Section
        row += 2
        ws[f'A{row}'] = "📋 ACTION ITEMS"
        ws[f'A{row}'].font = self.fonts['header']
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws.merge_cells(f'A{row}:F{row}')
        
        # Mock action items based on top opportunities
        top_opps = data['opportunities'].head(5)
        row += 1
        headers = ['Ticker', 'Action', 'Entry Level', 'Stop Level', 'Target', 'R:R']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i, value=header)
            cell.font = self.fonts['subheader']
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        
        row += 1
        for _, opp in top_opps.iterrows():
            ws[f'A{row}'] = opp['ticker']
            ws[f'B{row}'] = opp['setup_stage']
            ws[f'C{row}'] = f"${opp['entry_level']:.2f}"
            ws[f'D{row}'] = f"${opp['stop_level']:.2f}"
            ws[f'E{row}'] = f"${opp['target_level']:.2f}"
            ws[f'F{row}'] = f"{opp['risk_reward']:.1f}:1"
            
            # Color code by setup stage
            if opp['setup_stage'] == 'Entry':
                action_color = self.colors['strong_bullish']
            elif opp['setup_stage'] == 'Watch':
                action_color = self.colors['neutral']
            else:
                action_color = self.colors['background']
            
            ws[f'B{row}'].fill = PatternFill(start_color=action_color, end_color=action_color, fill_type='solid')
            row += 1
    
    def create_sector_chart_tab(self, ws, data):
        """Create sector performance visualization tab"""
        
        ws.title = "Sector Analysis"
        
        # Title
        ws['A1'] = "📊 SECTOR PERFORMANCE & ROTATION"
        ws['A1'].font = self.fonts['title']
        ws.merge_cells('A1:G1')
        
        # Sector performance table with chart
        sector_df = data['sector_performance'].sort_values('daily_change_pct', ascending=False)
        
        # Headers
        headers = ['Sector ETF', 'Sector Name', 'Daily %', 'Weekly %', 'Monthly %', 'Trend', 'Signal']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=header)
            cell.font = self.fonts['subheader']
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        
        # Data with conditional formatting
        for i, (_, row) in enumerate(sector_df.iterrows(), 4):
            ws[f'A{i}'] = row['etf']
            ws[f'B{i}'] = row['sector_name']
            ws[f'C{i}'] = f"{row['daily_change_pct']:+.1f}%"
            ws[f'D{i}'] = f"{row['weekly_change_pct']:+.1f}%"
            ws[f'E{i}'] = f"{row['monthly_change_pct']:+.1f}%"
            ws[f'F{i}'] = row['trend_status']
            
            # Signal emoji based on performance
            if row['daily_change_pct'] > 1.0:
                ws[f'G{i}'] = '🔥'
            elif row['daily_change_pct'] > 0:
                ws[f'G{i}'] = '📈'
            elif row['daily_change_pct'] > -1.0:
                ws[f'G{i}'] = '📉'
            else:
                ws[f'G{i}'] = '🔻'
            
            # Color code performance
            for col in ['C', 'D', 'E']:
                value = row[f"{['daily', 'weekly', 'monthly'][ord(col)-67]}_change_pct"]
                if value > 0:
                    color = self.colors['bullish']
                elif value < -1:
                    color = self.colors['bearish']
                else:
                    color = self.colors['neutral']
                ws[f'{col}{i}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    
    def create_complete_dashboard(self, scenario_name="default"):
        """Create complete Excel dashboard with all tabs"""
        
        print(f"🔨 Building Excel dashboard for scenario: {scenario_name}")
        
        # Load data
        data = self.load_mock_data()
        if not data:
            print("❌ Failed to load mock data")
            return None
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create tabs
        market_pulse_ws = wb.create_sheet("Market Pulse")
        screener_ws = wb.create_sheet("Screener Heatmap")
        alerts_ws = wb.create_sheet("Alerts & Actions")
        sector_ws = wb.create_sheet("Sector Analysis")
        
        # Build each tab
        self.create_market_pulse_tab(market_pulse_ws, data)
        self.create_screener_heatmap_tab(screener_ws, data)
        self.create_alerts_tab(alerts_ws, data)
        self.create_sector_chart_tab(sector_ws, data)
        
        # Save workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"trading_dashboard_mock_{scenario_name}_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        wb.save(filepath)
        
        print(f"✅ Dashboard created: {filepath}")
        print(f"  • Tabs: {len(wb.sheetnames)}")
        print(f"  • Market signal: {data['market_pulse']['overall_signal']}")
        print(f"  • Top opportunities: {len(data['opportunities'])}")
        
        return str(filepath)

def build_all_scenario_dashboards():
    """Build dashboards for all scenarios"""
    
    scenarios = ['bullish_market', 'bearish_market', 'neutral_market', 'volatile_market']
    
    for scenario in scenarios:
        print(f"\n📊 Building dashboard for {scenario}...")
        
        # Set data directory for this scenario
        data_dir = Path(f"sample_data/{scenario}")
        if not data_dir.exists():
            print(f"⚠️ Scenario data not found: {data_dir}")
            continue
        
        builder = TradingDashboardBuilder(data_dir=data_dir)
        dashboard_path = builder.create_complete_dashboard(scenario_name=scenario)
        
        if dashboard_path:
            print(f"✅ {scenario} dashboard: {Path(dashboard_path).name}")

if __name__ == "__main__":
    print("🚀 Trading Dashboard Builder")
    print("=" * 40)
    
    # Build default dashboard
    builder = TradingDashboardBuilder()
    dashboard_path = builder.create_complete_dashboard()
    
    # Build scenario dashboards if they exist
    print("\n🎬 Building scenario dashboards...")
    build_all_scenario_dashboards()
    
    print("\n✅ Dashboard building completed!")
    print("Check the 'output/' directory for generated Excel files.")