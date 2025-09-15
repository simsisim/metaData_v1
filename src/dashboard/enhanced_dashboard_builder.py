#!/usr/bin/env python3
"""
Enhanced Excel Dashboard Builder for Trading System
===================================================

Professional-grade Excel dashboard inspired by Koyfin's design principles.
Features improved visual hierarchy, better color scheme, and enhanced layout.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import json
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference, LineChart, ScatterChart, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

class EnhancedTradingDashboard:
    """Enhanced professional Excel dashboard builder inspired by Koyfin design"""
    
    def __init__(self, data_dir="sample_data", output_dir="output"):
        self.data_dir = Path(data_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # Enhanced color scheme (Koyfin-inspired professional palette)
        self.colors = {
            # Primary signals (more subtle, professional)
            'bullish_primary': 'FF00C851',      # Material Green
            'bearish_primary': 'FFFF4444',      # Material Red
            'neutral_primary': 'FFFFBB33',      # Material Amber
            'bullish_light': 'FFC8E6C9',        # Light Green
            'bearish_light': 'FFFFCDD2',        # Light Red
            'neutral_light': 'FFFFF8E1',        # Light Amber
            
            # Headers and structure (professional grays)
            'header_dark': 'FF263238',          # Blue Gray 900
            'header_medium': 'FF37474F',        # Blue Gray 800
            'header_light': 'FF546E7A',         # Blue Gray 600
            'subheader': 'FF78909C',            # Blue Gray 500
            
            # Background and borders
            'background_primary': 'FFFAFAFA',   # Gray 50
            'background_secondary': 'FFF5F5F5', # Gray 100
            'border_light': 'FFE0E0E0',         # Gray 300
            'border_medium': 'FFBDBDBD',        # Gray 400
            
            # Data emphasis
            'accent_blue': 'FF2196F3',          # Material Blue
            'accent_orange': 'FFFF9800',        # Material Orange
            'accent_purple': 'FF9C27B0',        # Material Purple
            
            # Status indicators
            'success': 'FF4CAF50',              # Material Green
            'warning': 'FFFF9800',              # Material Orange
            'danger': 'FFF44336',               # Material Red
            'info': 'FF2196F3'                  # Material Blue
        }
        
        # Enhanced typography (Koyfin-inspired)
        self.fonts = {
            'title': Font(name='Arial', size=18, bold=True, color='FFFFFF'),
            'header_primary': Font(name='Arial', size=14, bold=True, color='FFFFFF'),
            'header_secondary': Font(name='Arial', size=12, bold=True, color='FFFFFF'),
            'subheader': Font(name='Arial', size=11, bold=True, color='FF37474F'),
            'data_primary': Font(name='Arial', size=10, color='FF212121'),
            'data_secondary': Font(name='Arial', size=9, color='FF424242'),
            'signal_strong': Font(name='Arial', size=11, bold=True),
            'metric_large': Font(name='Arial', size=12, bold=True),
            'metric_small': Font(name='Arial', size=9)
        }
        
        # Enhanced borders
        self.borders = {
            'thin': Border(
                left=Side(style='thin', color='FFE0E0E0'),
                right=Side(style='thin', color='FFE0E0E0'),
                top=Side(style='thin', color='FFE0E0E0'),
                bottom=Side(style='thin', color='FFE0E0E0')
            ),
            'medium': Border(
                left=Side(style='medium', color='FFBDBDBD'),
                right=Side(style='medium', color='FFBDBDBD'),
                top=Side(style='medium', color='FFBDBDBD'),
                bottom=Side(style='medium', color='FFBDBDBD')
            ),
            'header': Border(
                bottom=Side(style='thick', color='FF37474F')
            )
        }
    
    def load_mock_data(self):
        """Load all mock data files"""
        try:
            # Load market pulse data
            with open(self.data_dir / 'market_pulse_data.json', 'r') as f:
                market_pulse = json.load(f)
            
            # Load CSV files
            screener_results = pd.read_csv(self.data_dir / 'screener_results.csv')
            sector_performance = pd.read_csv(self.data_dir / 'sector_performance.csv')
            index_technical = pd.read_csv(self.data_dir / 'index_technical.csv')
            alerts = pd.read_csv(self.data_dir / 'alerts_data.csv')
            opportunities = pd.read_csv(self.data_dir / 'top_opportunities.csv')
            
            return {
                'market_pulse': market_pulse,
                'screener_results': screener_results,
                'sector_performance': sector_performance,
                'index_technical': index_technical,
                'alerts': alerts,
                'opportunities': opportunities
            }
            
        except Exception as e:
            print(f"❌ Error loading mock data: {e}")
            return None
    
    def apply_cell_style(self, cell, style_type):
        """Apply consistent styling to cells"""
        styles = {
            'title': {
                'font': self.fonts['title'],
                'fill': PatternFill(start_color=self.colors['header_dark'], end_color=self.colors['header_dark'], fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'header_primary': {
                'font': self.fonts['header_primary'],
                'fill': PatternFill(start_color=self.colors['header_dark'], end_color=self.colors['header_dark'], fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': self.borders['header']
            },
            'header_secondary': {
                'font': self.fonts['header_secondary'],
                'fill': PatternFill(start_color=self.colors['header_medium'], end_color=self.colors['header_medium'], fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'subheader': {
                'font': self.fonts['subheader'],
                'fill': PatternFill(start_color=self.colors['background_secondary'], end_color=self.colors['background_secondary'], fill_type='solid'),
                'alignment': Alignment(horizontal='left', vertical='center'),
                'border': self.borders['thin']
            },
            'data_primary': {
                'font': self.fonts['data_primary'],
                'alignment': Alignment(horizontal='left', vertical='center'),
                'border': self.borders['thin']
            },
            'data_center': {
                'font': self.fonts['data_primary'],
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': self.borders['thin']
            },
            'metric_positive': {
                'font': self.fonts['signal_strong'],
                'fill': PatternFill(start_color=self.colors['bullish_light'], end_color=self.colors['bullish_light'], fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': self.borders['thin']
            },
            'metric_negative': {
                'font': self.fonts['signal_strong'],
                'fill': PatternFill(start_color=self.colors['bearish_light'], end_color=self.colors['bearish_light'], fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': self.borders['thin']
            },
            'metric_neutral': {
                'font': self.fonts['signal_strong'],
                'fill': PatternFill(start_color=self.colors['neutral_light'], end_color=self.colors['neutral_light'], fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': self.borders['thin']
            }
        }
        
        if style_type in styles:
            style = styles[style_type]
            for attr, value in style.items():
                setattr(cell, attr, value)
    
    def create_enhanced_market_pulse_tab(self, ws, data):
        """Create enhanced Market Pulse tab with Koyfin-inspired design"""
        
        ws.title = "📊 Market Pulse"
        
        # Enhanced title section
        ws['A1'] = "MARKET PULSE DASHBOARD"
        self.apply_cell_style(ws['A1'], 'title')
        ws.merge_cells('A1:L1')
        ws.row_dimensions[1].height = 30
        
        # Subtitle with timestamp
        ws['A2'] = f"Live Market Analysis • {datetime.now().strftime('%A, %B %d, %Y • %H:%M')}"
        ws['A2'].font = Font(name='Arial', size=11, italic=True, color='FF546E7A')
        ws.merge_cells('A2:L2')
        
        # Market Status Cards (Row 4-8)
        self._create_status_cards(ws, data, start_row=4)
        
        # GMI Analysis Section (Row 10-15)
        self._create_gmi_section(ws, data, start_row=10)
        
        # Risk Monitor Section (Row 10-15, columns G-L)
        self._create_risk_monitor_section(ws, data, start_row=10, start_col=7)
        
        # Market Breadth Section (Row 17-22)
        self._create_breadth_section(ws, data, start_row=17)
        
        # Sector Performance with Enhanced Visualization (Row 24-35)
        self._create_enhanced_sector_section(ws, data, start_row=24)
        
        # Add performance charts (Row 37+)
        self._add_sector_performance_chart(ws, data, start_row=37)
        
        # Apply professional styling
        self._apply_worksheet_formatting(ws)
    
    def _create_status_cards(self, ws, data, start_row):
        """Create enhanced status cards with better visual design"""
        
        market_signal = data['market_pulse']['overall_signal']
        confidence = data['market_pulse']['confidence']
        
        # Overall Market Status Card
        ws[f'A{start_row}'] = "MARKET STATUS"
        self.apply_cell_style(ws[f'A{start_row}'], 'header_primary')
        ws.merge_cells(f'A{start_row}:D{start_row}')
        
        # Signal display with enhanced styling
        signal_emoji = {'BULLISH': '🟢', 'BEARISH': '🔴', 'NEUTRAL': '🟡'}[market_signal]
        ws[f'A{start_row+1}'] = f"{signal_emoji} {market_signal}"
        ws[f'A{start_row+1}'].font = Font(name='Arial', size=16, bold=True, color='FF212121')
        
        ws[f'A{start_row+2}'] = f"Confidence: {confidence}"
        ws[f'A{start_row+2}'].font = self.fonts['data_primary']
        
        # Risk Level Card
        dd_data = data['market_pulse']['distribution_days']
        risk_level = 'HIGH' if dd_data['warning_level'] == 'SEVERE' else 'MEDIUM' if dd_data['warning_level'] == 'CAUTION' else 'LOW'
        
        ws[f'F{start_row}'] = "RISK ASSESSMENT"
        self.apply_cell_style(ws[f'F{start_row}'], 'header_primary')
        ws.merge_cells(f'F{start_row}:I{start_row}')
        
        ws[f'F{start_row+1}'] = f"Risk Level: {risk_level}"
        risk_colors = {'LOW': 'metric_positive', 'MEDIUM': 'metric_neutral', 'HIGH': 'metric_negative'}
        self.apply_cell_style(ws[f'F{start_row+1}'], risk_colors[risk_level])
        
        ws[f'F{start_row+2}'] = f"Distribution Days: {dd_data['count']}/25"
        ws[f'F{start_row+2}'].font = self.fonts['data_primary']
        
        # Market Health Card
        breadth = data['market_pulse']['market_breadth']
        ws[f'J{start_row}'] = "MARKET BREADTH"
        self.apply_cell_style(ws[f'J{start_row}'], 'header_primary')
        ws.merge_cells(f'J{start_row}:L{start_row}')
        
        net_highs = breadth['net_highs']
        ws[f'J{start_row+1}'] = f"Net H/L: {net_highs:+d}"
        breadth_style = 'metric_positive' if net_highs > 50 else 'metric_negative' if net_highs < -50 else 'metric_neutral'
        self.apply_cell_style(ws[f'J{start_row+1}'], breadth_style)
        
        ws[f'J{start_row+2}'] = f"Universe: {breadth['universe_size']} stocks"
        ws[f'J{start_row+2}'].font = self.fonts['data_secondary']
    
    def _create_gmi_section(self, ws, data, start_row):
        """Create enhanced GMI analysis section"""
        
        ws[f'A{start_row}'] = "GMI ANALYSIS"
        self.apply_cell_style(ws[f'A{start_row}'], 'header_secondary')
        ws.merge_cells(f'A{start_row}:F{start_row}')
        
        # GMI table headers
        headers = ['Index', 'Signal', 'Score', 'Trend', 'Momentum', 'Status']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=i, value=header)
            self.apply_cell_style(cell, 'subheader')
        
        # GMI data with enhanced visualization
        gmi_data = data['market_pulse']['gmi_analysis']
        row = start_row + 2
        
        for index, gmi_info in gmi_data.items():
            score = gmi_info['gmi_score']
            signal = gmi_info['gmi_signal']
            components = gmi_info.get('components', {})
            
            # Index name
            ws[f'A{row}'] = index
            self.apply_cell_style(ws[f'A{row}'], 'data_primary')
            
            # Signal with enhanced styling
            signal_emoji = {'GREEN': '🟢', 'RED': '🔴', 'NEUTRAL': '🟡'}[signal]
            ws[f'B{row}'] = f"{signal_emoji} {signal}"
            signal_style = {'GREEN': 'metric_positive', 'RED': 'metric_negative', 'NEUTRAL': 'metric_neutral'}[signal]
            self.apply_cell_style(ws[f'B{row}'], signal_style)
            
            # Score with progress bar effect
            ws[f'C{row}'] = f"{score}/4"
            score_style = 'metric_positive' if score >= 3 else 'metric_negative' if score <= 1 else 'metric_neutral'
            self.apply_cell_style(ws[f'C{row}'], score_style)
            
            # Component indicators
            trend_status = "✓" if components.get('short_trend', False) else "✗"
            momentum_status = "✓" if components.get('momentum', False) else "✗"
            
            ws[f'D{row}'] = trend_status
            ws[f'E{row}'] = momentum_status
            
            # Overall status
            ws[f'F{row}'] = "STRONG" if score >= 3 else "WEAK" if score <= 1 else "MIXED"
            
            for col in ['D', 'E', 'F']:
                self.apply_cell_style(ws[f'{col}{row}'], 'data_center')
            
            row += 1
    
    def _create_risk_monitor_section(self, ws, data, start_row, start_col):
        """Create enhanced risk monitoring section"""
        
        col_letter = get_column_letter(start_col)
        end_col_letter = get_column_letter(start_col + 5)
        
        ws[f'{col_letter}{start_row}'] = "RISK MONITOR"
        self.apply_cell_style(ws[f'{col_letter}{start_row}'], 'header_secondary')
        ws.merge_cells(f'{col_letter}{start_row}:{end_col_letter}{start_row}')
        
        # Distribution Days Analysis
        dd_data = data['market_pulse']['distribution_days']
        
        risk_metrics = [
            ('Distribution Days', f"{dd_data['count']}/25", dd_data['warning_level']),
            ('Warning Level', dd_data['warning_level'], dd_data['warning_level']),
            ('Lookback Period', f"{dd_data['lookback_period']} days", 'NONE')
        ]
        
        # FTD Analysis
        ftd_data = data['market_pulse']['follow_through_days']
        if ftd_data:
            risk_metrics.append(('Latest FTD', ftd_data.get('date', 'N/A'), 'OPPORTUNITY'))
        else:
            risk_metrics.append(('Follow-Through', 'None detected', 'NONE'))
        
        # Create risk metrics table
        row = start_row + 1
        for metric_name, value, status in risk_metrics:
            ws[f'{col_letter}{row}'] = metric_name
            ws[f'{get_column_letter(start_col+1)}{row}'] = value
            
            # Status indicator
            status_emoji = {'SEVERE': '🚨', 'CAUTION': '⚠️', 'OPPORTUNITY': '🚀', 'NONE': '✅'}
            ws[f'{get_column_letter(start_col+2)}{row}'] = status_emoji.get(status, '•')
            
            # Apply styling based on status
            if status in ['SEVERE']:
                style = 'metric_negative'
            elif status in ['CAUTION']:
                style = 'metric_neutral'
            elif status in ['OPPORTUNITY']:
                style = 'metric_positive'
            else:
                style = 'data_primary'
            
            for c in range(start_col, start_col + 3):
                self.apply_cell_style(ws[f'{get_column_letter(c)}{row}'], style)
            
            row += 1
    
    def _create_breadth_section(self, ws, data, start_row):
        """Create enhanced market breadth section"""
        
        ws[f'A{start_row}'] = "MARKET BREADTH ANALYSIS"
        self.apply_cell_style(ws[f'A{start_row}'], 'header_secondary')
        ws.merge_cells(f'A{start_row}:L{start_row}')
        
        breadth = data['market_pulse']['market_breadth']
        
        # Breadth metrics with visual indicators
        breadth_metrics = [
            ('New 52-Week Highs', breadth['new_highs'], 'positive'),
            ('New 52-Week Lows', breadth['new_lows'], 'negative'),
            ('Net New Highs', breadth['net_highs'], 'net'),
            ('Universe Size', breadth['universe_size'], 'info')
        ]
        
        # Create horizontal metrics layout
        col = 1
        for metric_name, value, metric_type in breadth_metrics:
            # Metric name
            ws.cell(row=start_row+1, column=col, value=metric_name)
            self.apply_cell_style(ws.cell(row=start_row+1, column=col), 'subheader')
            
            # Metric value
            if metric_type == 'net':
                display_value = f"{value:+d}"
            else:
                display_value = f"{value:,}" if isinstance(value, int) else str(value)
            
            ws.cell(row=start_row+2, column=col, value=display_value)
            
            # Color coding
            if metric_type == 'positive' or (metric_type == 'net' and value > 50):
                style = 'metric_positive'
            elif metric_type == 'negative' or (metric_type == 'net' and value < -50):
                style = 'metric_negative'
            else:
                style = 'metric_neutral'
            
            self.apply_cell_style(ws.cell(row=start_row+2, column=col), style)
            
            col += 3  # Space between metrics
        
        # Breadth signal summary
        breadth_signal = breadth['breadth_signal']
        ws[f'A{start_row+4}'] = f"Market Breadth Signal: {breadth_signal}"
        signal_style = {'HEALTHY': 'metric_positive', 'UNHEALTHY': 'metric_negative', 'NEUTRAL': 'metric_neutral'}[breadth_signal]
        self.apply_cell_style(ws[f'A{start_row+4}'], signal_style)
        ws.merge_cells(f'A{start_row+4}:L{start_row+4}')
    
    def _create_enhanced_sector_section(self, ws, data, start_row):
        """Create enhanced sector performance section with charts"""
        
        ws[f'A{start_row}'] = "SECTOR PERFORMANCE & ROTATION"
        self.apply_cell_style(ws[f'A{start_row}'], 'header_secondary')
        ws.merge_cells(f'A{start_row}:L{start_row}')
        
        sector_df = data['sector_performance'].sort_values('daily_change_pct', ascending=False)
        
        # Enhanced table headers
        headers = ['Sector ETF', 'Sector Name', 'Today %', 'Week %', 'Month %', 'Volume', 'Trend', 'Signal']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=i, value=header)
            self.apply_cell_style(cell, 'subheader')
        
        # Sector data with enhanced formatting
        for i, (_, row) in enumerate(sector_df.iterrows(), start_row+2):
            # ETF symbol
            ws[f'A{i}'] = row['etf']
            self.apply_cell_style(ws[f'A{i}'], 'data_primary')
            
            # Sector name
            ws[f'B{i}'] = row['sector_name']
            self.apply_cell_style(ws[f'B{i}'], 'data_primary')
            
            # Performance percentages with enhanced color coding
            daily_pct = row['daily_change_pct']
            weekly_pct = row['weekly_change_pct']
            monthly_pct = row['monthly_change_pct']
            
            ws[f'C{i}'] = f"{daily_pct:+.1f}%"
            ws[f'D{i}'] = f"{weekly_pct:+.1f}%"
            ws[f'E{i}'] = f"{monthly_pct:+.1f}%"
            
            # Apply color gradient based on performance
            for col, pct in [('C', daily_pct), ('D', weekly_pct), ('E', monthly_pct)]:
                if pct > 1.5:
                    self.apply_cell_style(ws[f'{col}{i}'], 'metric_positive')
                elif pct < -1.5:
                    self.apply_cell_style(ws[f'{col}{i}'], 'metric_negative')
                elif pct > 0:
                    cell = ws[f'{col}{i}']
                    cell.fill = PatternFill(start_color='FFE8F5E8', end_color='FFE8F5E8', fill_type='solid')
                    cell.font = self.fonts['data_primary']
                    cell.border = self.borders['thin']
                else:
                    cell = ws[f'{col}{i}']
                    cell.fill = PatternFill(start_color='FFFCE4EC', end_color='FFFCE4EC', fill_type='solid')
                    cell.font = self.fonts['data_primary']
                    cell.border = self.borders['thin']
            
            # Volume
            ws[f'F{i}'] = f"{row['volume']:,.0f}"
            self.apply_cell_style(ws[f'F{i}'], 'data_center')
            
            # Trend status with color coding
            trend = row['trend_status']
            ws[f'G{i}'] = trend
            trend_colors = {
                'DARK_GREEN': 'metric_positive',
                'LIGHT_GREEN': PatternFill(start_color='FFE8F5E8', end_color='FFE8F5E8', fill_type='solid'),
                'YELLOW': 'metric_neutral',
                'RED': 'metric_negative'
            }
            
            if isinstance(trend_colors.get(trend), str):
                self.apply_cell_style(ws[f'G{i}'], trend_colors[trend])
            else:
                ws[f'G{i}'].fill = trend_colors.get(trend, PatternFill())
                ws[f'G{i}'].font = self.fonts['data_primary']
                ws[f'G{i}'].border = self.borders['thin']
            
            # Signal strength
            if daily_pct > 1.5:
                signal = '🔥'
            elif daily_pct > 0.5:
                signal = '📈'
            elif daily_pct > -0.5:
                signal = '➡️'
            else:
                signal = '📉'
            
            ws[f'H{i}'] = signal
            self.apply_cell_style(ws[f'H{i}'], 'data_center')
    
    def create_enhanced_opportunities_tab(self, ws, data):
        """Create enhanced opportunities tab with better visualization"""
        
        ws.title = "🚀 Opportunities"
        
        # Enhanced title
        ws['A1'] = "TOP TRADING OPPORTUNITIES"
        self.apply_cell_style(ws['A1'], 'title')
        ws.merge_cells('A1:K1')
        ws.row_dimensions[1].height = 30
        
        # Subtitle
        opportunities_count = len(data['opportunities'])
        ws['A2'] = f"Live Screener Results • {opportunities_count} Opportunities Identified"
        ws['A2'].font = Font(name='Arial', size=11, italic=True, color='FF546E7A')
        ws.merge_cells('A2:K2')
        
        # Enhanced opportunities table
        headers = ['Rank', 'Ticker', 'Signal', 'Screener', 'Score', 'Price', 'Entry', 'Stop', 'Target', 'R:R', 'Volume']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i, value=header)
            self.apply_cell_style(cell, 'subheader')
        
        # Top opportunities with enhanced formatting
        top_opps = data['opportunities'].head(15)
        
        for rank, (_, opp) in enumerate(top_opps.iterrows(), 1):
            row = 4 + rank
            
            # Rank
            ws[f'A{row}'] = rank
            self.apply_cell_style(ws[f'A{row}'], 'data_center')
            
            # Ticker
            ws[f'B{row}'] = opp['ticker']
            ws[f'B{row}'].font = Font(name='Arial', size=11, bold=True, color='FF212121')
            
            # Signal strength with enhanced styling
            signal = opp['signal_strength']
            signal_emojis = {'STRONG': '🟢', 'MODERATE': '🟡', 'WEAK': '🔴'}
            ws[f'C{row}'] = f"{signal_emojis[signal]} {signal}"
            
            signal_styles = {'STRONG': 'metric_positive', 'MODERATE': 'metric_neutral', 'WEAK': 'metric_negative'}
            self.apply_cell_style(ws[f'C{row}'], signal_styles[signal])
            
            # Screener name (cleaned)
            screener_name = opp['primary_screener'].replace('_', ' ').title()
            ws[f'D{row}'] = screener_name
            self.apply_cell_style(ws[f'D{row}'], 'data_primary')
            
            # Score with color gradient
            score = opp['score']
            ws[f'E{row}'] = f"{score:.1f}/10"
            if score >= 8.0:
                self.apply_cell_style(ws[f'E{row}'], 'metric_positive')
            elif score >= 6.0:
                ws[f'E{row}'].fill = PatternFill(start_color='FFFFF8E1', end_color='FFFFF8E1', fill_type='solid')
                ws[f'E{row}'].font = self.fonts['data_primary']
                ws[f'E{row}'].border = self.borders['thin']
            else:
                self.apply_cell_style(ws[f'E{row}'], 'metric_negative')
            
            # Price data
            ws[f'F{row}'] = f"${opp['current_price']:.2f}"
            ws[f'G{row}'] = f"${opp['entry_level']:.2f}"
            ws[f'H{row}'] = f"${opp['stop_level']:.2f}"
            ws[f'I{row}'] = f"${opp['target_level']:.2f}"
            
            for col in ['F', 'G', 'H', 'I']:
                self.apply_cell_style(ws[f'{col}{row}'], 'data_center')
            
            # Risk/Reward ratio
            rr = opp['risk_reward']
            ws[f'J{row}'] = f"{rr:.1f}:1"
            rr_style = 'metric_positive' if rr >= 2.0 else 'metric_neutral' if rr >= 1.5 else 'metric_negative'
            self.apply_cell_style(ws[f'J{row}'], rr_style)
            
            # Volume
            volume = opp['volume']
            if volume >= 1000000:
                vol_display = f"{volume/1000000:.1f}M"
            elif volume >= 1000:
                vol_display = f"{volume/1000:.0f}K"
            else:
                vol_display = str(volume)
            
            ws[f'K{row}'] = vol_display
            self.apply_cell_style(ws[f'K{row}'], 'data_center')
        
        # Add screener summary section
        summary_start_row = len(data['opportunities']) + 8  # Dynamic positioning
        self._create_screener_summary_section(ws, data, summary_start_row)
        
        # Add screener performance chart
        chart_start_row = summary_start_row + 12
        self._add_screener_performance_chart(ws, data, chart_start_row)
    
    def _create_screener_summary_section(self, ws, data, start_row):
        """Create screener suite summary with performance indicators"""
        
        ws[f'A{start_row}'] = "SCREENER SUITE PERFORMANCE"
        self.apply_cell_style(ws[f'A{start_row}'], 'header_secondary')
        ws.merge_cells(f'A{start_row}:F{start_row}')
        
        # Mock screener performance data
        screener_summary = [
            ('Stockbee Suite', 12, 85, 'STRONG'),
            ('Qullamaggie Suite', 6, 72, 'MODERATE'),
            ('Volume Suite', 8, 88, 'STRONG'),
            ('Gold Launch Pad', 3, 65, 'MODERATE'),
            ('RTI Screener', 7, 78, 'STRONG'),
            ('ADL Screener', 4, 58, 'WEAK'),
            ('Guppy GMMA', 5, 69, 'MODERATE'),
            ('ATR1 Suite', 9, 82, 'STRONG')
        ]
        
        headers = ['Screener Suite', 'Hits', 'Avg Score', 'Performance', 'Status', 'Trend']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=i, value=header)
            self.apply_cell_style(cell, 'subheader')
        
        for i, (screener, hits, avg_score, performance) in enumerate(screener_summary, start_row+2):
            ws[f'A{i}'] = screener
            ws[f'B{i}'] = hits
            ws[f'C{i}'] = f"{avg_score}/100"
            ws[f'D{i}'] = performance
            
            # Status indicator
            if performance == 'STRONG':
                ws[f'E{i}'] = '🟢'
                perf_style = 'metric_positive'
            elif performance == 'MODERATE':
                ws[f'E{i}'] = '🟡'
                perf_style = 'metric_neutral'
            else:
                ws[f'E{i}'] = '🔴'
                perf_style = 'metric_negative'
            
            # Trend indicator (mock)
            trend = '📈' if avg_score > 75 else '➡️' if avg_score > 60 else '📉'
            ws[f'F{i}'] = trend
            
            # Apply styling
            for col in ['A', 'B', 'C']:
                self.apply_cell_style(ws[f'{col}{i}'], 'data_primary')
            
            self.apply_cell_style(ws[f'D{i}'], perf_style)
            
            for col in ['E', 'F']:
                self.apply_cell_style(ws[f'{col}{i}'], 'data_center')
    
    def _add_sector_performance_chart(self, ws, data, start_row):
        """Add sector performance bar chart"""
        try:
            sector_df = data['sector_performance'].sort_values('daily_change_pct', ascending=False)
            
            # Create chart data section
            chart_start_row = start_row
            ws[f'A{chart_start_row}'] = "SECTOR PERFORMANCE CHART"
            self.apply_cell_style(ws[f'A{chart_start_row}'], 'header_secondary')
            ws.merge_cells(f'A{chart_start_row}:F{chart_start_row}')
            
            # Add chart data
            chart_data_row = chart_start_row + 2
            ws[f'A{chart_data_row}'] = "Sector"
            ws[f'B{chart_data_row}'] = "Daily %"
            
            for i, (_, row) in enumerate(sector_df.iterrows(), chart_data_row + 1):
                ws[f'A{i}'] = row['sector_name']
                ws[f'B{i}'] = row['daily_change_pct']
            
            # Create bar chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Sector Performance (Daily %)"
            chart.y_axis.title = "Performance %"
            chart.x_axis.title = "Sectors"
            
            # Set data references
            data_ref = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + len(sector_df))
            cats_ref = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + len(sector_df))
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            # Style the chart
            chart.width = 15
            chart.height = 8
            
            # Add chart to worksheet
            ws.add_chart(chart, f"H{chart_start_row}")
            
        except Exception as e:
            logger.error(f"Error creating sector performance chart: {e}")
    
    def _add_screener_performance_chart(self, ws, data, start_row):
        """Add screener suite performance visualization"""
        try:
            # Create screener performance data
            screener_data = [
                ('Stockbee Suite', 12, 85),
                ('Qullamaggie Suite', 6, 72),
                ('Volume Suite', 8, 88),
                ('Gold Launch Pad', 3, 65),
                ('RTI Screener', 7, 78),
                ('ADL Screener', 4, 58),
                ('Guppy GMMA', 5, 69),
                ('ATR1 Suite', 9, 82)
            ]
            
            # Chart title
            ws[f'A{start_row}'] = "SCREENER PERFORMANCE ANALYSIS"
            self.apply_cell_style(ws[f'A{start_row}'], 'header_secondary')
            ws.merge_cells(f'A{start_row}:F{start_row}')
            
            # Add chart data
            chart_data_row = start_row + 2
            headers = ['Screener', 'Hits', 'Avg Score']
            
            for i, header in enumerate(headers, 1):
                ws.cell(row=chart_data_row, column=i, value=header)
            
            for i, (screener, hits, score) in enumerate(screener_data, chart_data_row + 1):
                ws[f'A{i}'] = screener
                ws[f'B{i}'] = hits
                ws[f'C{i}'] = score
            
            # Create dual-axis chart (hits vs scores)
            chart = BarChart()
            chart.type = "col"
            chart.style = 12
            chart.title = "Screener Suite Performance"
            chart.y_axis.title = "Hits Count"
            
            # Data for hits (column B)
            hits_ref = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + len(screener_data))
            cats_ref = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + len(screener_data))
            
            chart.add_data(hits_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            # Style the chart
            chart.width = 12
            chart.height = 8
            
            # Add chart to worksheet
            ws.add_chart(chart, f"H{start_row}")
            
        except Exception as e:
            logger.error(f"Error creating screener performance chart: {e}")
    
    def _apply_worksheet_formatting(self, ws):
        """Apply overall worksheet formatting"""
        
        # Set default column widths
        default_widths = {
            'A': 15, 'B': 20, 'C': 12, 'D': 15, 'E': 12, 'F': 12,
            'G': 15, 'H': 12, 'I': 12, 'J': 10, 'K': 12, 'L': 15
        }
        
        for col, width in default_widths.items():
            ws.column_dimensions[col].width = width
        
        # Set default row height
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
        # Apply background to entire sheet
        for row in ws.iter_rows():
            for cell in row:
                if not cell.fill.start_color.index:  # Only if no fill applied
                    cell.fill = PatternFill(start_color=self.colors['background_primary'], 
                                          end_color=self.colors['background_primary'], fill_type='solid')
    
    def create_complete_dashboard(self, scenario_name="enhanced"):
        """Create complete enhanced dashboard with Koyfin-inspired design"""
        return self.create_enhanced_complete_dashboard(scenario_name)
    
    def create_enhanced_complete_dashboard(self, scenario_name="enhanced"):
        """Create complete enhanced dashboard with Koyfin-inspired design"""
        
        print(f"🎨 Building enhanced Excel dashboard: {scenario_name}")
        
        # Load data
        data = self.load_mock_data()
        if not data:
            print("❌ Failed to load mock data")
            return None
        
        # Create workbook with enhanced styling
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create enhanced tabs
        market_pulse_ws = wb.create_sheet("📊 Market Pulse")
        opportunities_ws = wb.create_sheet("🚀 Opportunities")
        portfolio_ws = wb.create_sheet("💼 Portfolio")
        alerts_ws = wb.create_sheet("🚨 Alerts")
        
        # Build enhanced tabs
        self.create_enhanced_market_pulse_tab(market_pulse_ws, data)
        self.create_enhanced_opportunities_tab(opportunities_ws, data)
        self._create_enhanced_portfolio_tab(portfolio_ws, data)
        self._create_enhanced_alerts_tab(alerts_ws, data)
        
        # Save enhanced workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"enhanced_trading_dashboard_{scenario_name}_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        wb.save(filepath)
        
        print(f"✅ Enhanced dashboard created: {filepath}")
        print(f"🎨 Design: Koyfin-inspired professional layout")
        print(f"📊 Features: Enhanced colors, typography, and visual hierarchy")
        
        return str(filepath)
    
    def _create_enhanced_alerts_tab(self, ws, data):
        """Create enhanced alerts tab with improved design"""
        
        ws.title = "🚨 Alerts"
        
        # Enhanced title
        ws['A1'] = "MARKET ALERTS & ACTION INTELLIGENCE"
        self.apply_cell_style(ws['A1'], 'title')
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 30
        
        # Critical alerts section
        ws['A3'] = "🚨 CRITICAL ALERTS"
        self.apply_cell_style(ws['A3'], 'header_primary')
        ws.merge_cells('A3:H3')
        
        # Enhanced alert categories with smart prioritization
        alert_categories = {
            'CRITICAL': {'color': 'danger', 'emoji': '🚨', 'title': 'CRITICAL - IMMEDIATE ACTION'},
            'HIGH': {'color': 'warning', 'emoji': '⚠️', 'title': 'HIGH PRIORITY'},
            'MEDIUM': {'color': 'info', 'emoji': '📊', 'title': 'MEDIUM PRIORITY'},
            'LOW': {'color': 'success', 'emoji': 'ℹ️', 'title': 'INFORMATIONAL'}
        }
        
        current_row = 5
        
        for priority, alert_config in alert_categories.items():
            priority_alerts = data['alerts'][data['alerts']['priority'] == priority]
            
            if not priority_alerts.empty:
                # Section header
                ws[f'A{current_row}'] = f"{alert_config['emoji']} {alert_config['title']} ALERTS"
                
                header_color = self.colors.get(alert_config['color'], self.colors['header_medium'])
                ws[f'A{current_row}'].fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
                ws[f'A{current_row}'].font = self.fonts['header_secondary']
                ws.merge_cells(f'A{current_row}:H{current_row}')
                
                current_row += 1
                
                # Alert items
                for _, alert in priority_alerts.iterrows():
                    ws[f'A{current_row}'] = f"• {alert['message']}"
                    ws[f'A{current_row}'].font = self.fonts['data_primary']
                    ws.merge_cells(f'A{current_row}:H{current_row}')
                    current_row += 1
                
                current_row += 1  # Space between sections
        
        # Action items section
        current_row += 1
        ws[f'A{current_row}'] = "📋 EXECUTION CHECKLIST"
        self.apply_cell_style(ws[f'A{current_row}'], 'header_secondary')
        ws.merge_cells(f'A{current_row}:H{current_row}')
        
        # Enhanced action items table
        action_headers = ['Ticker', 'Action', 'Entry Level', 'Stop Loss', 'Target', 'R:R', 'Setup Stage', 'Priority']
        current_row += 1
        
        for i, header in enumerate(action_headers, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            self.apply_cell_style(cell, 'subheader')
        
        # Action items from top opportunities
        top_actions = data['opportunities'].head(8)
        
        for _, action in top_actions.iterrows():
            current_row += 1
            
            ws[f'A{current_row}'] = action['ticker']
            ws[f'B{current_row}'] = action['setup_stage']
            ws[f'C{current_row}'] = f"${action['entry_level']:.2f}"
            ws[f'D{current_row}'] = f"${action['stop_level']:.2f}"
            ws[f'E{current_row}'] = f"${action['target_level']:.2f}"
            ws[f'F{current_row}'] = f"{action['risk_reward']:.1f}:1"
            ws[f'G{current_row}'] = action['setup_stage']
            
            # Priority based on signal strength
            priority = 'HIGH' if action['signal_strength'] == 'STRONG' else 'MEDIUM'
            ws[f'H{current_row}'] = priority
            
            # Apply styling
            for col in ['A', 'B', 'C', 'D', 'E', 'G']:
                self.apply_cell_style(ws[f'{col}{current_row}'], 'data_center')
            
            # Color code R:R ratio
            rr = action['risk_reward']
            rr_style = 'metric_positive' if rr >= 2.0 else 'metric_neutral' if rr >= 1.5 else 'metric_negative'
            self.apply_cell_style(ws[f'F{current_row}'], rr_style)
            
            # Color code priority
            priority_style = 'metric_positive' if priority == 'HIGH' else 'metric_neutral'
            self.apply_cell_style(ws[f'H{current_row}'], priority_style)
    
    def _create_enhanced_portfolio_tab(self, ws, data):
        """Create enhanced portfolio tracking tab"""
        
        ws.title = "💼 Portfolio"
        
        # Enhanced title
        ws['A1'] = "PORTFOLIO PERFORMANCE DASHBOARD"
        self.apply_cell_style(ws['A1'], 'title')
        ws.merge_cells('A1:L1')
        ws.row_dimensions[1].height = 30
        
        # Portfolio summary section
        ws['A3'] = "PORTFOLIO OVERVIEW"
        self.apply_cell_style(ws['A3'], 'header_primary')
        ws.merge_cells('A3:F3')
        
        # Mock portfolio summary metrics
        portfolio_metrics = [
            ('Total Value', '$125,450', 'metric_positive'),
            ('Daily P&L', '+$2,180 (+1.76%)', 'metric_positive'),
            ('Weekly P&L', '+$5,620 (+4.69%)', 'metric_positive'),
            ('Active Positions', '8 positions', 'data_primary')
        ]
        
        # Create portfolio metrics cards
        col = 1
        for metric_name, value, style in portfolio_metrics:
            ws.cell(row=4, column=col, value=metric_name)
            self.apply_cell_style(ws.cell(row=4, column=col), 'subheader')
            
            ws.cell(row=5, column=col, value=value)
            self.apply_cell_style(ws.cell(row=5, column=col), style)
            
            col += 3
        
        # Portfolio positions table
        ws['A7'] = "CURRENT POSITIONS"
        self.apply_cell_style(ws['A7'], 'header_secondary')
        ws.merge_cells('A7:L7')
        
        # Portfolio table headers
        portfolio_headers = ['Ticker', 'Shares', 'Entry Price', 'Current Price', 'P&L $', 'P&L %', 'Days Held', 'Signal', 'Action', 'Risk Level']
        for i, header in enumerate(portfolio_headers, 1):
            cell = ws.cell(row=8, column=i, value=header)
            self.apply_cell_style(cell, 'subheader')
        
        # Sample portfolio data (would come from real portfolio in production)
        portfolio_positions = [
            ('AAPL', 100, 175.50, 185.25, 975, 5.55, 12, 'HOLD', 'Monitor', 'LOW'),
            ('NVDA', 25, 420.00, 445.75, 643.75, 6.13, 8, 'STRONG', 'Hold', 'MEDIUM'),
            ('TSLA', 50, 245.80, 240.15, -282.50, -2.30, 18, 'WEAK', 'Review', 'HIGH'),
            ('META', 30, 298.50, 315.20, 501, 5.59, 6, 'MODERATE', 'Hold', 'LOW'),
            ('AMZN', 15, 145.30, 152.85, 113.25, 5.19, 22, 'HOLD', 'Monitor', 'MEDIUM')
        ]
        
        for i, (ticker, shares, entry, current, pnl_dollar, pnl_pct, days, signal, action, risk) in enumerate(portfolio_positions, 9):
            ws[f'A{i}'] = ticker
            ws[f'B{i}'] = shares
            ws[f'C{i}'] = f"${entry:.2f}"
            ws[f'D{i}'] = f"${current:.2f}"
            ws[f'E{i}'] = f"${pnl_dollar:.2f}"
            ws[f'F{i}'] = f"{pnl_pct:+.1f}%"
            ws[f'G{i}'] = days
            ws[f'H{i}'] = signal
            ws[f'I{i}'] = action
            ws[f'J{i}'] = risk
            
            # Apply styling based on P&L
            if pnl_pct > 0:
                pnl_style = 'metric_positive'
            elif pnl_pct < -2:
                pnl_style = 'metric_negative'
            else:
                pnl_style = 'metric_neutral'
            
            # Apply base styling
            for col in ['A', 'B', 'C', 'D', 'G', 'H', 'I', 'J']:
                self.apply_cell_style(ws[f'{col}{i}'], 'data_center')
            
            # Special styling for P&L columns
            self.apply_cell_style(ws[f'E{i}'], pnl_style)
            self.apply_cell_style(ws[f'F{i}'], pnl_style)
            
            # Risk level color coding
            risk_colors = {'LOW': 'metric_positive', 'MEDIUM': 'metric_neutral', 'HIGH': 'metric_negative'}
            self.apply_cell_style(ws[f'J{i}'], risk_colors.get(risk, 'data_center'))
        
        # Portfolio allocation chart section
        ws['A15'] = "PORTFOLIO ALLOCATION"
        self.apply_cell_style(ws['A15'], 'header_secondary')
        ws.merge_cells('A15:F15')
        
        # Allocation data
        allocation_data = [
            ('Technology', 45.2, 'XLK'),
            ('Consumer Disc.', 23.8, 'XLY'),
            ('Communication', 18.5, 'XLC'),
            ('Industrials', 12.5, 'XLI')
        ]
        
        # Create allocation table
        ws['A16'] = "Sector"
        ws['B16'] = "Weight %"
        ws['C16'] = "ETF Proxy"
        
        for i, header_cell in enumerate(['A16', 'B16', 'C16']):
            self.apply_cell_style(ws[header_cell], 'subheader')
        
        for i, (sector, weight, etf) in enumerate(allocation_data, 17):
            ws[f'A{i}'] = sector
            ws[f'B{i}'] = f"{weight:.1f}%"
            ws[f'C{i}'] = etf
            
            # Color code based on weight
            if weight > 30:
                weight_style = 'metric_positive'
            elif weight > 15:
                weight_style = 'metric_neutral'
            else:
                weight_style = 'data_center'
            
            self.apply_cell_style(ws[f'A{i}'], 'data_primary')
            self.apply_cell_style(ws[f'B{i}'], weight_style)
            self.apply_cell_style(ws[f'C{i}'], 'data_center')


def create_enhanced_dashboard_from_real_data(config, user_config, results_dir, timeframe='daily', data_reader=None):
    """
    Create enhanced dashboard using real trading system data
    
    Args:
        config: System configuration
        user_config: User configuration
        results_dir: Directory containing real screener results
        timeframe: Data timeframe
        data_reader: DataReader instance
        
    Returns:
        Path to generated enhanced dashboard
    """
    try:
        from .real_data_connector import RealDataConnector
        
        print("🎨 Creating enhanced dashboard with real data...")
        
        # Create real data connector
        connector = RealDataConnector(config, user_config, data_reader)
        
        # Generate real dashboard data
        dashboard_data = connector.generate_real_dashboard_data(results_dir, timeframe)
        
        # Create enhanced dashboard output directory
        enhanced_output_dir = Path(results_dir) / 'dashboards' / 'enhanced'
        enhanced_output_dir.mkdir(parents=True, exist_ok=True)
        
        # Create temporary data directory for enhanced dashboard builder
        temp_data_dir = Path(results_dir) / 'dashboard_temp_enhanced'
        temp_data_dir.mkdir(exist_ok=True)
        
        # Save dashboard data in format expected by enhanced builder
        _save_enhanced_dashboard_input_files(dashboard_data, temp_data_dir)
        
        # Build enhanced dashboard
        enhanced_builder = EnhancedTradingDashboard(data_dir=temp_data_dir, output_dir=enhanced_output_dir)
        dashboard_path = enhanced_builder.create_enhanced_complete_dashboard(scenario_name='real_data')
        
        print(f"✅ Enhanced dashboard created: {dashboard_path}")
        return dashboard_path
        
    except Exception as e:
        print(f"❌ Error creating enhanced dashboard: {e}")
        raise


def _save_enhanced_dashboard_input_files(dashboard_data, temp_dir):
    """Save dashboard data for enhanced builder"""
    
    # Save market pulse as JSON
    with open(temp_dir / 'market_pulse_data.json', 'w') as f:
        json.dump(dashboard_data['market_pulse'], f, indent=2, default=str)
    
    # Save DataFrames as CSV
    csv_files = ['screener_results', 'sector_performance', 'index_technical', 'alerts', 'opportunities']
    
    for file_key in csv_files:
        if file_key in dashboard_data and isinstance(dashboard_data[file_key], pd.DataFrame):
            dashboard_data[file_key].to_csv(temp_dir / f'{file_key}.csv', index=False)
    
    # Create enhanced summary
    summary = {
        'generation_timestamp': datetime.now().isoformat(),
        'data_source': 'real_trading_system_enhanced',
        'market_signal': dashboard_data['market_pulse']['overall_signal'],
        'total_screener_hits': len(dashboard_data.get('screener_results', [])),
        'sectors_analyzed': len(dashboard_data.get('sector_performance', [])),
        'indexes_analyzed': len(dashboard_data.get('index_technical', [])),
        'dashboard_version': 'enhanced_v2'
    }
    
    with open(temp_dir / 'data_summary.json', 'w') as f:
        json.dump(summary, f, indent=2)


if __name__ == "__main__":
    print("🎨 Enhanced Trading Dashboard Builder")
    print("=" * 45)
    
    # Build enhanced dashboard with mock data
    enhanced_builder = EnhancedTradingDashboard(data_dir="../mock_dashboard/sample_data")
    dashboard_path = enhanced_builder.create_enhanced_complete_dashboard()
    
    print(f"\n✅ Enhanced dashboard building completed!")
    print(f"📁 Check output directory: {enhanced_builder.output_dir}")